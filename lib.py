# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------
#
#    Copyright (C) 2017  jeo Software  (http://www.jeosoft.com.ar)
#    All Rights Reserved.
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
# -----------------------------------------------------------------------------------

# -*- coding: utf-8 -*-
import openpyxl


class Product(object):
    def __init__(self, data):
        self._pack = False
        self._desc = ' '
        self._ctrl = data[0]
        if data[1]:
            self._desc = data[1].replace('\n', '')
            self._pack = self._desc.find('PACK') > -1
        self._code = self._check_code(str(data[2]))
        self._price = data[3] if data[3] else 0.0

    def _check_code(self, code):

        return code

    @property
    def pack(self):
        return self._pack

    @property
    def ctrl(self):
        return self._ctrl

    @property
    def code(self):
        return self._code

    @property
    def desc(self):
        return self._desc

    @property
    def price(self):
        return self._price

    def list(self):
        return u'{:10}  ${:6.2f}  {} {:10}  '.format(self.code, self.price, 'PACK' if self._pack else '', self.desc)


class MilaWorksheet(object):
    def __init__(self, filename):

        CTRL = 4
        DESC = 6
        CODE = 8
        PRICE = 10

        def decode(acell):
            ret = acell.value
            return ret

        self._prods = []

        wb = openpyxl.load_workbook(filename=filename, read_only=True)
        sheet = wb.get_sheet_by_name('PEDIDO COT MM-MP')

        # itero sobre toda la planilla por filas
        for row in sheet.iter_rows(min_row=23, min_col=4, max_col=10, max_row=627):
            rowlist = []
            # itero en toda la fila por celdas
            for cell in row:
                # considero solo algunas celdas
                if cell.column in [CTRL, DESC, CODE, PRICE]:
                    if cell.value == '9999-99':
                        print 'Lectura de planilla mila ok'
                        return
                    rowlist.append(decode(cell))
            # me quedo solo con las rows marcadas MM o MP
            if rowlist[0] in ['MM', 'MP', 'M20A', 'P20A']:
                self._prods.append(Product(rowlist))

        """
        sheet = wb.get_sheet_by_name('PEDIDO COT MM-MP')

        # itero sobre la otra hoja de la planilla por filas
        for row in sheet.iter_rows(min_row=1, min_col=4, max_col=11, max_row=35):
            rowlist = []
            # itero en toda la fila por celdas
            for cell in row:
                # considero solo algunas celdas
                if cell.column in [CTRL, DESC, CODE, PRICE]:
                    rowlist.append(decode(cell))
            # me quedo solo con las rows marcadas MM o MP
            if rowlist[0] in ['MM', 'MP']:
                self._prods.append(Product(rowlist))
        """
    def list(self):
        return self._prods

    def prod(self, default_code):
        """ Search a product in mila worksheets """
        p = False
        for p in self._prods:
            if p.code == default_code:
                return p

