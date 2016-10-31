# -*- coding: utf-8 -*-
# -----------------------------------------------------------------------------------
#
#    Copyright (C) 2016  jeo Software  (http://www.jeo-soft.com.ar)
#    All Rights Reserved.
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#    thanks to https://automatetheboringstuff.com/ where inspiration comes from
#    http://openpyxl.readthedocs.org/
#
# -----------------------------------------------------------------------------------
import openpyxl

wb = openpyxl.load_workbook(filename='mila.xlsx')
sheet = wb.get_sheet_by_name('PEDIDO COT MM-MP')

print '"code","cant","price","desc"'
for row in range(23, 1000):
    desc = sheet.cell(row=row, column=6).value
    code = sheet.cell(row=row, column=8).value
    cant = sheet.cell(row=row, column=9).value
    price = sheet.cell(row=row, column=10).value
    desc = desc.replace('\n', ' ') if desc else False
    if price and cant:
        print u'"{}","{}","{}","{}"'.format(code, cant, price,desc)


# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
